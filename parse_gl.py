#!/usr/bin/env python3
"""Parse all GL tabs from finance's Software Spend workbook (Oct 2025 - Jul 2026)."""
import json, re, datetime
import openpyxl
from collections import defaultdict

NM = json.load(open('naming-map.json'))
ALIASES = [(re.compile(p, re.I), app) for p, app in NM['vendor_aliases']]
INCLUDED = set(NM['included_accounts'])
EXCL_ACCT = set(NM['non_saas_exclusions']['accounts'])

TABS = {  # tab name -> (year, month-key)
    'October Spend': ('2025','Oct25'), 'November Spend': ('2025','Nov25'), 'December Spend': ('2025','Dec25'),
    'Jan 26 spend': ('2026','Jan'), 'Feb 26 spend': ('2026','Feb'), 'Mar 26 spend': ('2026','Mar'),
    'Apr 26 Spend': ('2026','Apr'), 'May 26 Spend': ('2026','May'), 'JUNE 26 Spend': ('2026','Jun'),
    'July 26 Spend': ('2026','Jul'),
}
ACCT_CANDIDATES = ['account full name','full name','account','distribution account','account #','split']
FIVE = re.compile(r'\b(\d{5})\b')

def find_header(ws):
    for r in range(1, min(6, ws.max_row)+1):
        vals = {str(ws.cell(row=r,column=c).value).strip().lower(): c
                for c in range(1, ws.max_column+1) if ws.cell(row=r,column=c).value is not None}
        if ('transaction date' in vals or 'date' in vals) and ('amount' in vals or 'debit' in vals):
            return r, vals
    raise RuntimeError(f'no header row in {ws.title}')

def resolve(vendor):
    for rx, app in ALIASES:
        if rx.search(vendor):
            return app
    return None

def last5(acct_str):
    if acct_str is None: return None
    s = str(acct_str)
    m = FIVE.findall(s)
    return int(m[-1]) if m else None

rows = []
activity25 = defaultdict(lambda: defaultdict(float))  # app -> {Oct25/Nov25/Dec25: amt}
unresolved = defaultdict(float)
for tab,(yr,mk) in TABS.items():
    wb = openpyxl.load_workbook('inputs/software_spend.xlsx', data_only=True)
    ws = wb[tab]
    hr, cols = find_header(ws)
    def cget(rr, names):
        for n in names:
            if n in cols:
                return ws.cell(row=rr, column=cols[n]).value
        return None
    n_inc = n_exc = 0
    for r in range(hr+1, ws.max_row+1):
        date = cget(r, ['transaction date','date'])
        amt  = cget(r, ['amount'])
        if amt is None:
            d = cget(r,['debit']); c = cget(r,['credit'])
            if d is None and c is None: continue
            amt = (float(d or 0)) - (float(c or 0))
        try: amt = float(amt)
        except (TypeError, ValueError): continue
        if date is None and amt == 0: continue
        if date is None: continue  # totals/summary rows carry no date
        typ  = str(cget(r,['transaction type']) or '').strip()
        num  = str(cget(r,['num']) or '').strip()
        name = str(cget(r,['name']) or '').strip()
        desc = str(cget(r,['description','memo/description','memo']) or '').strip()
        if name in ('0.0','0'): name = ''
        acct_raw = None
        for cand in ['account full name','full name','distribution account','account','account #']:
            v = cget(r,[cand])
            if v is not None and str(v).strip() not in ('','0','0.0'):
                acct_raw = v
                if last5(v) is not None: break
        acct = last5(acct_raw)
        vendor = (name + ' ' + desc).strip().upper()
        if not vendor: continue
        app = resolve(vendor)
        # exclusions
        if app == '__EXCLUDE__' or acct in EXCL_ACCT:
            n_exc += 1; continue
        # inclusion: known account, or account-less/other rows that resolve to a known app
        if acct in INCLUDED or (app is not None and (acct is None or typ == 'Journal Entry')):
            pass
        else:
            n_exc += 1; continue
        if app == 'Salesforce - Slack Enterprise' and abs(amt) < 5000:
            app = 'Salesforce - Slack Internal'   # runbook rule: sub-$5K slack rows are the internal plan
        if app is None:
            unresolved[f'{mk}: {vendor[:60]}'] += amt
            app = '__UNRESOLVED__'
        n_inc += 1
        if yr == '2025':
            activity25[app][mk] += amt
        else:
            rows.append(dict(mon=mk, app=app, amt=round(amt,2), acct=acct, typ=typ, num=num, vendor=vendor[:80]))
    print(f'{tab:<18} included={n_inc:4d} excluded={n_exc:4d}')

json.dump(rows, open('inputs/gl_rows.json','w'))
json.dump({k:dict(v) for k,v in activity25.items()}, open('inputs/activity_2025.json','w'), indent=1)
print(f'\n2026 rows: {len(rows)}')
if unresolved:
    print('\n!! UNRESOLVED VENDORS (must fix before shipping):')
    for k,v in sorted(unresolved.items(), key=lambda x:-abs(x[1])):
        print(f'   {v:>10,.2f}  {k}')
else:
    print('All vendors resolved.')

months = ['Jan','Feb','Mar','Apr','May','Jun','Jul']
tot = defaultdict(float)
for r in rows: tot[r['mon']] += r['amt']
print('\nRaw monthly totals (pre-classification):')
base = {'Jan':151655.48,'Feb':157003.65,'Mar':163215.17,'Apr':168080.11,'May':146127.40}
for m in months:
    b = base.get(m)
    diff = f'  diff vs baseline: {tot[m]-b:+,.2f}' if b else ''
    print(f'  {m}: {tot[m]:>12,.2f}{diff}')
